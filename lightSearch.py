import os
import pandas as pd
import re
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlencode

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook

import base64
import shutil
from tqdm import trange
from img2pdf import conpdf
import time
import random

headers = {
    # 'Accept-Language': 'zh-CN,zh;q=0.9',
    # 'Cache-Control': 'max-age=0',
    # 'Connection': 'keep-alive',
    # 'Sec-Fetch-Dest': 'document',
    # 'Sec-Fetch-Mode': 'navigate',
    # 'Sec-Fetch-Site': 'none',
    # 'Sec-Fetch-User': '?1',
    # 'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1'
}
folder_path = "./tableDatas"
columns = 6  # 列
rows = 1  # 行
myFormatList = ["", "DOC", "PDF", "PPT"]


def get_user_input(prompt, validation_func):
    """
    该函数用于接收用户的输入并验证是否符合要求。
    参数：
    - prompt: str，表示要显示给用户的提示信息。
    - validation_func: function，表示用于验证用户输入的函数。
    返回值：
    - 用户输入的字符串。
    """
    while True:
        user_input = input(prompt)
        if validation_func(user_input):
            return user_input
        else:
            print("输入有误，请重新输入。")


def validate_yes_no(user_input):
    return user_input.lower() == 'y' or user_input.lower() == 'n' or user_input.lower() == ''


def validate_file_format(user_input):
    return user_input.isdigit() and 0 <= int(user_input) <= 3


def validate_search_result(user_input):
    return user_input.isdigit() and 1 <= int(user_input) <= 10


def remove_keywords(text):
    keywords = ['docx', 'pdf', 'ppt', '_ppt', 'word', 'doc']
    for keyword in keywords:
        text = text.replace(keyword.lower(), '').replace(keyword.upper(), '')
    return text


def process_filename(output_dir, title):
    to_remove = [" - 道客巴巴", "精品", "精选", "ppt", "doc", "pdf", "word", "_", "[]", " "]
    output_title = title
    for s in to_remove:
        output_title = output_title.replace(s, "")
    tepfile = f"{output_dir}/{output_title}.pdf"
    if os.path.exists(tepfile):
        output_title += "内容"
    if len(output_title) < 5:
        output_title += "内容介绍"
    return output_title


def download(url, file_name_without):
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('log-level=3')
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(options=options)

    title = "output"

    try:
        driver.set_page_load_timeout(15)
        driver.get(url)
        title = driver.title
    except:
        print("Timeout - start download anyway.")

    # newtitle=f"{title}{timestamp}"
    print(f'道客巴巴: 《{title}》')
    # print(f'《{title.split(":")[1].strip()}》')
    # 睡眠5秒等待页面加载
    time.sleep(5)

    try:
        # 展开全部
        elem_cont_button = driver.find_element(By.ID, "continueButton")
        driver.execute_script("arguments[0].scrollIntoView(true);", elem_cont_button)
        actions = ActionChains(driver)
        actions.move_to_element(elem_cont_button).perform()
        time.sleep(0.5)
        elem_cont_button.click()
    except NoSuchElementException:
        pass

    # 获取页数
    num_of_pages = driver.find_element(By.ID, 'toolbar').find_element(By.ID, 'item-page-panel'). \
        find_element(By.CLASS_NAME, 'text').text
    num_of_pages = int(num_of_pages.split(' ')[-1])

    for i in range(5):
        # 缩放
        driver.find_element(By.ID, 'zoomInButton').click()
        time.sleep(0.5)

    if os.path.exists(f'./temp/{title}'):
            shutil.rmtree(f'./temp/{title}')
    os.makedirs(f'./temp/{title}')

    for pages in trange(num_of_pages):
        if pages >= 15:
            break
        time.sleep(0.5)
        canvas_id = "outer_page_" + str(pages + 1)
        pagepb_id = "page_" + str(pages + 1)

        try:
            element = driver.find_element(By.ID, canvas_id)
        except:
            time.sleep(1)
            element = driver.find_element(By.ID, canvas_id)

        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        actions = ActionChains(driver)
        actions.move_to_element(element).perform()
        time.sleep(0.5)
        # 执行js代码
        js_cmd = "var canvas = document.getElementById('{}');".format(pagepb_id) + \
                 "return canvas.toDataURL();"
        img_data = driver.execute_script(js_cmd)
        img_data = (img_data[22:]).encode()

        with open(f"./temp/{title}/{pages}.png", "wb") as fh:
            fh.write(base64.decodebytes(img_data))
    driver.quit()
    print('下载完毕，正在转码')

    # output_dir = './output'
    # if not os.path.exists(output_dir):
    #     os.makedirs(output_dir)

    # 要创建的目录路径
    dir_path = f'./output/{file_name_without}'
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
        print(f"已创建目录 {dir_path}")

    newOpTitle = process_filename(dir_path, title)
    output_file = f'output/{file_name_without}/{newOpTitle}.pdf'
    conpdf(output_file, f'temp/{title}', '.png')

    return newOpTitle


def getUrl(itemTitle, format, reItem):
    # url = "https://www.doc88.com/search/post.do?
    # f=0&h=1&t=1651200664&pageRange=1&pageNum=3&format="+myFormatList[format]+"&p=1&q="+itemTitle
    # 从 0 到 9 的数字中随机选择六个，生成一个六位随机数
    random_number = ''.join(random.sample('0123456789', 6))
    suslo = f"1698392{random_number}"
    base_url = "https://www.doc88.com/search/post.do"
    params = {
        'f': '0',
        'h': '1',
        't': '1651200664',
        'pageRange': '1',
        'pageNum': '3',
        'format': myFormatList[format],
        'p': '1',
        'q': itemTitle,
        '_': suslo
    }
    query_string = urlencode(params)
    url = f"{base_url}?{query_string}"
    print("搜索链接："+url)
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    divs = soup.find_all('div', class_='sd-list-con')

    if len(divs) >= 1:
        second_div = divs[reItem]
        link = second_div.find('a')['href']
        title = second_div.find('a').get('title')

        print("文章链接:", link)
        print("文章标题:", title)

        sd_intro = second_div.find('div', class_='sd-intro')

        if sd_intro:
            for span in sd_intro.find_all('span'):
                span.extract()
            for spacer in sd_intro.find_all('b', class_='spacer'):
                # 替换 <b class="spacer"></b> 为 -
                spacer.replace_with('-')

            # 去除空格和换行符并输出
            content = sd_intro.text.strip()
            content_without_spaces = re.sub(r'\s+', '', content)
            print("内容:", content_without_spaces)
        else:
            print("找不到对应的 class='sd-intro' 的 div")

        print("=================一个标题搜索完毕===================")
        return link
    else:
        # print("找不到div")
        return None


def getForms(format, reItem):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print("未检测到指定tableDatas文件夹，已创建成功，程序将退出！")
        exit()
    table_files = [file_name for file_name in os.listdir(folder_path) if file_name.endswith(('.xlsx', '.csv', '.xls'))]
    if not table_files:
        print("tableDatas文件夹中没有表格文件，程序将退出！")
        exit()

    # 遍历tableDatas文件夹中的表格文件
    for file_name in table_files:
        file_path = os.path.join(folder_path, file_name)
        file_name_without_ext = os.path.splitext(file_name)[0]  # 获取文件名并去除后缀

        setCellNumber = 2
        try:
            # 使用pandas读取表格文件。header参数设为None表示从第0行开始算起
            df = pd.read_excel(file_path, header=None)
            num_cols = len(df.columns)  # 获取表格列数
            if num_cols <= columns:
                print("表格列数不足，请检查文件：", file_path)
                continue

            column_G = df.iloc[rows:, columns]  # G列的索引为6
            if column_G.empty:
                print("该文件指定列没有数据")

            else:
                for celldata in column_G:

                    if pd.isna(celldata):
                        print('该单元格为空！')
                        continue
                    print("原标题：" + celldata)
                    print("搜索标题：" + remove_keywords(celldata))
                    reUrl = getUrl(remove_keywords(celldata), format, reItem)

                    if reUrl == None:
                        print("搜索结果似乎为空！")
                    else:
                        # download(reUrl, file_name_without_ext)
                        refinFile = download(reUrl, file_name_without_ext)
                        print(refinFile)
                        # 加载工作簿
                        wb = load_workbook(file_path)
                        # 选择活动工作表
                        ws = wb.active
                        # 在第二行的第一个单元格（B2）插入数据
                        ws.cell(row=setCellNumber, column=5).value = refinFile
                        # 保存工作簿
                        wb.save(file_path)
                        setCellNumber += 1

            # 创建目标文件夹（如果不存在）
            os.makedirs(f"output/{file_name_without_ext}", exist_ok=True)
            # 将源文件剪切到目标文件夹
            shutil.move(file_path, f"output/{file_name_without_ext}")
            print('=====================一份文件搜索完毕========================')

        except Exception as e:
            print(f"读取文件 {file_path} 失败: {str(e)}")


is_quick_start = get_user_input("是否启用默认方案(y/n) 直接回车可快速启动：", validate_yes_no)

if is_quick_start.lower() == 'y' or is_quick_start == '':
    print("执行默认方案...")
    getForms(1, 0)
else:
    file_format = get_user_input("获取 0:所有  1:doc   2:pdf  3:ppt 的文件(请输入序号)：", validate_file_format)

    search_result = get_user_input("获取搜索结果的第几项文件(1-10)：", validate_search_result)

    print(f"获取 {myFormatList[int(file_format)]} 格式的第 {search_result} 项搜索结果...")
    getForms(int(file_format), int(search_result))
    print("")
